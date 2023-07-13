import os
import csv
import pandas as pd
import re
import time 
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.drawing.image import Image
from openpyxl import load_workbook

ruta = '/Users/datos-lc/Documents/Colecciones ongoing/Empresas sancionadas - México/'

archivos = os.listdir(ruta)

for i in archivos:
    if i.endswith(".xlsx"):
        excel = load_workbook(os.path.join(ruta, i))
        ws = excel.active
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 50
        for column in ws.iter_cols(min_col=1, max_col=2, max_row=17):
            for cell in column:
                if cell.value is not None and isinstance(cell.value, str):
                    cell.alignment = Alignment(wrapText=True)

        
        pdf_name = i.split(".xlsx")[0]
        pdf_filename = os.path.join(ruta, pdf_name + ".pdf")
        excel.save(os.path.join(ruta, pdf_name + ".xlsx"))
        os.system(f'unoconv -f pdf "{pdf_filename}"')
